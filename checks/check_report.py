#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
LazyAWS - Report Builder (Excel)

- Reads RawData/*.json produced by other modules (S3/IAM/Lambda/EC2/Exposure/etc.)
- Builds one Excel file (Reports/LazyAWS_Report.xlsx) with separate sheets
- Status column shows severity (Critical/High/Medium/Low/OK/N/A)
- CLI request/response pulled from Aux.api_trace (or synthesized)
"""

from __future__ import annotations

import argparse
import glob
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# prefer package-relative import
try:
    from .report_severity import infer_check_code, severity_for, get_outcome, strip_ansi
except Exception:
    from report_severity import infer_check_code, severity_for, get_outcome, strip_ansi  # type: ignore

# -------------------- utils --------------------

def iso_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")

def ensure_dirs():
    Path("Reports").mkdir(parents=True, exist_ok=True)
    Path("RawData").mkdir(parents=True, exist_ok=True)

def load_json_files() -> List[Tuple[str, Dict[str, Any]]]:
    files = sorted(glob.glob("RawData/*.json"))
    out: List[Tuple[str, Dict[str, Any]]] = []
    for p in files:
        try:
            with open(p, "r", encoding="utf-8") as f:
                out.append((p, json.load(f)))
        except Exception:
            continue
    return out

def get_findings(doc: Dict[str, Any]) -> List[Dict[str, Any]]:
    # try common keys first
    for key in ("Findings", "findings", "Rows", "rows", "Checks", "checks"):
        v = doc.get(key)
        if isinstance(v, list):
            return v
    # nested under Report/Data
    for key in ("Report", "Data", "data"):
        v = doc.get(key)
        if isinstance(v, dict):
            for k2 in ("Findings", "findings", "Rows", "rows", "Checks", "checks"):
                v2 = v.get(k2)
                if isinstance(v2, list):
                    return v2
    return []

def group_by_service(items: List[Tuple[str, Dict[str, Any]]]) -> Dict[str, List[Tuple[str, Dict[str, Any]]]]:
    groups: Dict[str, List[Tuple[str, Dict[str, Any]]]] = {}
    for path, data in items:
        svc = (data.get("Meta", {}) or {}).get("Service") or infer_service_from_filename(path)
        groups.setdefault(str(svc), []).append((path, data))
    return groups

def infer_service_from_filename(path: str) -> str:
    name = Path(path).name.lower()
    if name.startswith("s3_"): return "S3"
    if name.startswith("iam_"): return "IAM"
    if name.startswith("lambda_"): return "Lambda"
    if name.startswith("ec2_"): return "EC2"
    if name.startswith("exposure_"): return "Exposure"
    return "Unknown"

def extract_api_list_from_check(check_text: str) -> List[str]:
    if not check_text:
        return []
    api_line = None
    for line in str(check_text).splitlines():
        if line.strip().startswith("API:"):
            api_line = line
    if not api_line:
        return []
    raw = api_line.split("API:", 1)[1].strip()
    if not raw or raw.upper() == "N/A":
        return []
    toks = [t.strip() for t in raw.split("/") if t.strip()]
    norm = []
    for t in toks:
        t = re.sub(r"\s+", "", t)
        if ":" in t:
            svc, op = t.split(":", 1)
            norm.append(f"{svc.lower()}:{op}")
        else:
            norm.append(t.lower())
    return norm

def api_trace_from_json(doc: Dict[str, Any]) -> List[Dict[str, Any]]:
    aux = doc.get("Aux", {}) or {}
    # canonical
    if isinstance(aux.get("api_trace"), list):
        return aux["api_trace"]  # type: ignore
    # legacy
    if isinstance(aux.get("CliTrace"), list):
        return aux["CliTrace"]  # type: ignore
    return []

def to_kebab(op_name: str) -> str:
    s = re.sub(r'(?<!^)(?=[A-Z])', '-', op_name).lower()
    return s.replace('--', '-')

def render_cli_from_trace_item(item: Dict[str, Any]) -> Tuple[str, str]:
    """
    Return (cli_request, cli_response) strings for a trace item.
    Prefer human-ready 'aws_cli'/'aws_cli_output' if present.
    Also support 'request_str'/'response_str' overrides.
    Else synthesize from service/operation/request.
    """
    # explicit text fields first
    if item.get("request_str") or item.get("response_str"):
        return (
            strip_ansi(str(item.get("request_str", "")).strip()),
            strip_ansi(str(item.get("response_str", "")).strip()),
        )
    # preformatted aws_cli if present
    if item.get("aws_cli") or item.get("aws_cli_output"):
        return (
            strip_ansi(str(item.get("aws_cli", "")).strip()),
            strip_ansi(str(item.get("aws_cli_output", "")).strip()),
        )

    service = str(item.get("service", "")).lower()
    operation = str(item.get("operation", "")).strip()
    params = item.get("request", {}) or item.get("params", {}) or {}

    if service and operation:
        op_kebab = to_kebab(operation)
        req = f"aws {service} {op_kebab} --cli-input-json '{json.dumps(params, separators=(',',':'))}'"
    else:
        req = json.dumps(params)

    resp = item.get("response")
    if resp is None:
        out = ""
    else:
        out = json.dumps(resp, ensure_ascii=False, separators=(",", ":"), default=str)
        if len(out) > 4000:
            out = out[:4000] + "... [truncated]"
    return (req, out)

def filter_trace_by_api(check_text: str, api_trace: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    wanted = set(extract_api_list_from_check(check_text))
    if not wanted:
        return api_trace
    out = []
    for it in api_trace:
        svc = str(it.get("service", "")).lower()
        op = str(it.get("operation", ""))
        if f"{svc}:{op}".lower() in wanted:
            out.append(it)
    return out or api_trace

def join_multiline(items: List[str], limit: int = 10) -> str:
    if not items:
        return ""
    sliced = items[:limit]
    text = "\n\n".join(sliced)
    if len(items) > limit:
        text += f"\n\n... {len(items)-limit} more ..."
    return text

SEV_COLORS = {
    "Critical": "FF3D3D","High":"FF9900","Medium":"FFC000","Low":"92D050","OK":"A9D08E","N/A":"D0CECE",
}

def style_sheet(ws):
    ws.freeze_panes = "A2"
    hdr_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {1:45, 2:12, 3:60, 4:50, 5:60, 6:60}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def apply_severity_fill(cell, severity: str):
    color = SEV_COLORS.get(severity)
    if color:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(wrap_text=True, vertical="top")

# -------------------- sheet builder --------------------

def build_sheet(wb: Workbook, title: str, docs: List[Tuple[str, Dict[str, Any]]]) -> int:
    ws = wb.create_sheet(title)
    ws.append(["Check", "Status", "Details", "Recommendation", "CLI request(s)", "CLI response(s)"])
    style_sheet(ws)
    rows_written = 0

    for _, doc in docs:
        findings = get_findings(doc)
        trace = api_trace_from_json(doc)
        if not isinstance(findings, list):
            continue

        for f in findings:
            outcome = get_outcome(f)  # normalized OK/WARN/BAD/N/A
            if outcome in ("OK","N/A","INFO",""):
                continue  # show only failed/problematic

            check_text = f.get("Check", "") or f.get("check", "")
            code = infer_check_code(check_text)
            severity = severity_for(code, f)

            filtered = filter_trace_by_api(check_text, trace)
            reqs, ress = [], []
            for it in filtered:
                r_cli, r_out = render_cli_from_trace_item(it)
                if r_cli: reqs.append(r_cli)
                if r_out: ress.append(r_out)

            ws.append([
                check_text,
                severity,
                strip_ansi(f.get("Details", "") or f.get("details","")),
                strip_ansi(f.get("Recommendation", "") or f.get("recommendation","")),
                join_multiline(reqs, 8),
                join_multiline(ress, 8),
            ])
            apply_severity_fill(ws.cell(row=ws.max_row, column=2), severity)
            for c in range(1, 7):
                ws.cell(row=ws.max_row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            rows_written += 1

    if rows_written == 0:
        ws.append(["(no failed findings in source data)", "OK", "", "", "", ""])
        apply_severity_fill(ws.cell(row=ws.max_row, column=2), "OK")
    return rows_written

# -------------------- main/report --------------------

def analyze(args):
    ensure_dirs()
    out_path = getattr(args, "out", None) or "Reports/LazyAWS_Report.xlsx"
    items = load_json_files()
    groups = group_by_service(items)

    wb = Workbook()
    wb.remove(wb.active)

    order = ["S3","IAM","Lambda","EC2","Exposure","Unknown"]
    counts = {}

    for svc in order:
        docs = groups.get(svc, [])
        if not docs: 
            continue
        rows = build_sheet(wb, svc, docs)
        counts[svc] = {"files": len(docs), "rows": rows}
    # any other services not in order
    for svc, docs in groups.items():
        if svc in order or not docs:
            continue
        rows = build_sheet(wb, svc, docs)
        counts[svc] = {"files": len(docs), "rows": rows}

    wb.save(out_path)
    for svc, st in counts.items():
        print(f"[{svc}] files={st['files']} rows={st['rows']}")
    print(f"[+] Saved Excel report to: {out_path}")

    meta = {"Service":"Report","Target":"Excel","Region":"—","Profile":getattr(args,"profile",None) or "default","TimeUTC":iso_now()}
    findings = [{"Check":"REPORT-001 Build Excel\nScope: Local\nAPI: N/A","Status":"OK","StatusPlain":"OK","Details":f"Saved to {Path(out_path).name}","Recommendation":"N/A"}]
    aux = {"output": out_path, "counts": counts}
    return meta, findings, aux

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="LazyAWS - Build Excel Report from RawData/")
    ap.add_argument("--out", default="Reports/LazyAWS_Report.xlsx", help="Output XLSX path")
    ap.add_argument("--profile", default=None, help="(for header only)")
    ap.add_argument("--region", default=None, help="(unused; header only)")
    args = ap.parse_args()
    try:
        from aws_common import run_check
        run_check(analyze, args)
    except Exception:
        meta, findings, aux = analyze(args)
        print(json.dumps({"Meta": meta, "Findings": findings, "Aux": aux}, indent=2))
