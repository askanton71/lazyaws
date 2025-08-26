#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
LazyAWS Report Builder
Reads JSON artifacts from RawData/ and generates an Excel report with sheets:
  - S3
  - EC2
  - IAM
  - LAMBDA

Each sheet summarizes targets and failed findings. For S3, each bucket section
includes: what was tested, positive vs negative expectation, remediation plan,
and CLI-style request+response (from api_trace).
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional, Tuple

# -------- Excel backend --------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except Exception as e:
    print("Missing dependency: openpyxl. Install it with:\n  pip install openpyxl", file=sys.stderr)
    raise

# -------- Helpers --------

def load_json(p: Path) -> Optional[Dict[str, Any]]:
    try:
        with p.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def list_raw_files(raw_dir: Path) -> List[Path]:
    if not raw_dir.exists():
        return []
    return sorted([p for p in raw_dir.glob("*.json") if p.is_file()])

def newest_for_target(files: List[Path]) -> Dict[Tuple[str, str], Path]:
    """
    Return a map (service_lower, target_lower) -> newest file path.
    Handles both stable (<service>_<target>.json) and timestamped files
    (<service>_<target>_YYYYmmdd-HHMMSS.json). Chooses newest by mtime.
    """
    best: Dict[Tuple[str, str], Path] = {}
    for p in files:
        data = load_json(p)
        if not data or "meta" not in data:
            continue
        meta = data["meta"]
        svc = str(meta.get("Service", "")).strip().lower()
        tgt = str(meta.get("Target", "")).strip().lower()
        if not svc or not tgt:
            continue
        key = (svc, tgt)
        if key not in best:
            best[key] = p
        else:
            if p.stat().st_mtime > best[key].stat().st_mtime:
                best[key] = p
    return best

def parse_check_label(check: str) -> Dict[str, str]:
    """
    Our 'Check' field is multi-line, e.g.:
      S3-006 Default Encryption (Bucket SSE)
      Scope: Bucket
      API: s3:GetBucketEncryption
    """
    lines = (check or "").splitlines()
    code = title = scope = api = ""
    if lines:
        first = lines[0].strip()
        if " " in first:
            code, title = first.split(" ", 1)
        else:
            code, title = first, ""
    for ln in lines[1:]:
        s = ln.strip()
        if s.lower().startswith("scope:"):
            scope = s.split(":", 1)[1].strip()
        elif s.lower().startswith("api:"):
            api = s.split(":", 1)[1].strip()
    return {"code": code, "title": title, "scope": scope, "api": api}

def pascal_to_snake(op: str) -> str:
    out = []
    for ch in op:
        if ch.isupper() and out:
            out.append("_")
        out.append(ch.lower())
    return "".join(out)

def api_from_check_api(api_field: str) -> Tuple[str, str]:
    """
    "s3:GetBucketEncryption" -> ("s3", "get_bucket_encryption")
    """
    if ":" in api_field:
        svc, op = api_field.split(":", 1)
        return svc.strip().lower(), pascal_to_snake(op.strip())
    return "", pascal_to_snake(api_field.strip())

def match_trace_for_api(api_trace: List[Dict[str, Any]], svc: str, op_snake: str) -> Optional[Dict[str, Any]]:
    """
    Heuristic: pick the LAST trace entry whose service == svc (or 's3control' for s3control)
    and api == op_snake. If not found, return None.
    """
    svc = svc.strip().lower()
    for entry in reversed(api_trace or []):
        if str(entry.get("service", "")).strip().lower() == svc and str(entry.get("api", "")).strip().lower() == op_snake:
            return entry
    return None

def ws_set_col_widths(ws, widths: List[int]):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

def apply_header_style(ws, cell_range: str):
    header_fill = PatternFill("solid", fgColor="222222")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="444444")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws[cell_range]:
        for c in row:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = border

def apply_table_borders(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = border

# -------- Knowledge base for S3 checks (what/positive/negative/remediation) --------

S3_CHECK_INFO: Dict[str, Dict[str, str]] = {
    "S3-001": {
        "what": "Reachability of the bucket (s3:HeadBucket). Validates name/region/permissions.",
        "positive": "HeadBucket succeeds.",
        "negative": "HeadBucket fails (NotFound/AccessDenied/etc).",
        "remed": "Verify bucket name and region; ensure your role has s3:HeadBucket and VPC endpoints if required."
    },
    "S3-002": {
        "what": "Bucket-level Public Access Block configuration.",
        "positive": "All four flags TRUE: BlockPublicAcls, IgnorePublicAcls, BlockPublicPolicy, RestrictPublicBuckets.",
        "negative": "Any flag is FALSE or config missing.",
        "remed": (
            "CLI:\n"
            "aws s3api put-public-access-block --bucket <bucket> --public-access-block-configuration "
            "BlockPublicAcls=true,IgnorePublicAcls=true,BlockPublicPolicy=true,RestrictPublicBuckets=true"
        )
    },
    "S3-003": {
        "what": "Account-level Public Access Block (s3control).",
        "positive": "All four flags TRUE account-wide.",
        "negative": "Config not set or any flag FALSE.",
        "remed": (
            "CLI:\n"
            "aws s3control put-public-access-block --account-id <ACCOUNT_ID> --public-access-block-configuration "
            "BlockPublicAcls=true,IgnorePublicAcls=true,BlockPublicPolicy=true,RestrictPublicBuckets=true"
        )
    },
    "S3-004": {
        "what": "OwnershipControls (ObjectOwnership) to disable ACLs.",
        "positive": "BucketOwnerEnforced.",
        "negative": "Not set or other mode.",
        "remed": (
            "CLI:\n"
            "aws s3api put-bucket-ownership-controls --bucket <bucket> --ownership-controls "
            "'{\"Rules\":[{\"ObjectOwnership\":\"BucketOwnerEnforced\"}]}'"
        )
    },
    "S3-005": {
        "what": "Bucket ACL grants that expose data (AllUsers/AuthenticatedUsers).",
        "positive": "Only owner; no public grants.",
        "negative": "Any public grants present.",
        "remed": (
            "Remove grants. CLI example to set private ACL:\n"
            "aws s3api put-bucket-acl --bucket <bucket> --acl private"
        )
    },
    "S3-006": {
        "what": "Default bucket encryption (SSE).",
        "positive": "aws:kms with CMK (optionally BucketKeyEnabled true).",
        "negative": "AES256 or not configured.",
        "remed": (
            "CLI (replace KMS key):\n"
            "aws s3api put-bucket-encryption --bucket <bucket> --server-side-encryption-configuration "
            "'{\"Rules\":[{\"ApplyServerSideEncryptionByDefault\":{\"SSEAlgorithm\":\"aws:kms\",\"KMSMasterKeyID\":\"<key-arn>\"},"
            "\"BucketKeyEnabled\":true}]}'"
        )
    },
    "S3-007": {
        "what": "Versioning state.",
        "positive": "Enabled.",
        "negative": "Suspended or NotConfigured.",
        "remed": "aws s3api put-bucket-versioning --bucket <bucket> --versioning-configuration Status=Enabled"
    },
    "S3-008": {
        "what": "Lifecycle configuration for transition/expiration.",
        "positive": "Rules defined reflecting retention/archival needs.",
        "negative": "No lifecycle rules.",
        "remed": (
            "Create lifecycle rules as needed. Example:\n"
            "aws s3api put-bucket-lifecycle-configuration --bucket <bucket> --lifecycle-configuration '{\n"
            "  \"Rules\": [\n"
            "    {\"ID\":\"archive\",\"Status\":\"Enabled\",\"Transitions\":[{\"Days\":30,\"StorageClass\":\"STANDARD_IA\"}],\n"
            "     \"Expiration\":{\"Days\":365},\"AbortIncompleteMultipartUpload\":{\"DaysAfterInitiation\":7}}\n"
            "  ]\n"
            "}'"
        )
    },
    "S3-009": {
        "what": "AbortIncompleteMultipartUpload lifecycle action.",
        "positive": "AbortIncompleteMultipartUpload present (e.g., 7 days).",
        "negative": "Missing AbortIncompleteMultipartUpload.",
        "remed": (
            "Add AbortIncompleteMultipartUpload to lifecycle rule (see S3-008 example)."
        )
    },
    "S3-010": {
        "what": "Server access logging.",
        "positive": "LoggingEnabled to a dedicated log bucket.",
        "negative": "Disabled.",
        "remed": (
            "Prepare a log bucket (versioned+encrypted), then:\n"
            "aws s3api put-bucket-logging --bucket <bucket> --bucket-logging-status "
            "'{\"LoggingEnabled\":{\"TargetBucket\":\"<log-bucket>\",\"TargetPrefix\":\"<bucket>/\"}}'"
        )
    },
    "S3-011": {
        "what": "Bucket policy denies non-TLS requests (aws:SecureTransport=false).",
        "positive": "Explicit Deny present.",
        "negative": "No TLS-enforcement deny.",
        "remed": (
            "Add a deny statement:\n"
            "{\n"
            "  \"Sid\":\"DenyInsecureTransport\",\"Effect\":\"Deny\",\"Principal\":\"*\",\"Action\":\"s3:*\",\n"
            "  \"Resource\":[\"arn:aws:s3:::<bucket>\",\"arn:aws:s3:::<bucket>/*\"],\n"
            "  \"Condition\":{\"Bool\":{\"aws:SecureTransport\":\"false\"}}\n"
            "}"
        )
    },
    "S3-012": {
        "what": "Bucket policy denies PutObject without aws:kms encryption.",
        "positive": "Explicit Deny for non-KMS uploads.",
        "negative": "No deny or allow of unencrypted/AES256 uploads.",
        "remed": (
            "Add denies (either StringNotEquals or Null conditions):\n"
            "{\n"
            "  \"Sid\":\"DenyUnencryptedObjectUploads\",\"Effect\":\"Deny\",\"Principal\":\"*\",\"Action\":\"s3:PutObject\",\n"
            "  \"Resource\":\"arn:aws:s3:::<bucket>/*\",\n"
            "  \"Condition\":{\"StringNotEquals\":{\"s3:x-amz-server-side-encryption\":\"aws:kms\"}}\n"
            "}\n"
            "and/or\n"
            "{\n"
            "  \"Sid\":\"DenyMissingSSE\",\"Effect\":\"Deny\",\"Principal\":\"*\",\"Action\":\"s3:PutObject\",\n"
            "  \"Resource\":\"arn:aws:s3:::<bucket>/*\",\n"
            "  \"Condition\":{\"Null\":{\"s3:x-amz-server-side-encryption\":\"true\"}}\n"
            "}"
        )
    },
    "S3-013": {
        "what": "Bucket policy JSON validity.",
        "positive": "Policy parses correctly.",
        "negative": "JSON parse errors.",
        "remed": "Validate JSON with a linter; use aws s3api put-bucket-policy with validated policy; test in IAM Policy Simulator."
    },
    "S3-014": {
        "what": "Presence of a bucket policy.",
        "positive": "Policy exists with least-privilege and required denies.",
        "negative": "No policy.",
        "remed": "Create a baseline policy including S3-011 and S3-012 and least-privilege principals."
    },
    "S3-015": {
        "what": "Static website hosting configuration.",
        "positive": "Disabled unless explicitly needed.",
        "negative": "Enabled.",
        "remed": "Disable website hosting: aws s3api delete-bucket-website --bucket <bucket>"
    },
    "S3-016": {
        "what": "CORS rules with wildcard origins.",
        "positive": "AllowedOrigins are specific domains.",
        "negative": "Wildcard '*' present.",
        "remed": (
            "Replace '*' with exact origins. CLI example:\n"
            "aws s3api put-bucket-cors --bucket <bucket> --cors-configuration '{\"CORSRules\":[{\"AllowedMethods\":[\"GET\"],"
            "\"AllowedOrigins\":[\"https://app.example.com\"],\"AllowedHeaders\":[\"*\"]}]}'"
        )
    },
    "S3-017": {
        "what": "Replication configuration (CRR/SRR).",
        "positive": "Replication rules exist if business requires.",
        "negative": "No replication though required by RPO.",
        "remed": "Configure replication with IAM role, KMS permissions, and destination bucket; use console wizard or s3api put-bucket-replication."
    },
    "S3-018": {
        "what": "Event notifications (SNS/SQS/Lambda).",
        "positive": "Configured with encrypted targets and strict access policies.",
        "negative": "Missing when required by workflows.",
        "remed": "Create notifications to required targets; ensure target encryption (KMS) and limited resource policies."
    },
    "S3-019": {
        "what": "S3 Access Points count/policy.",
        "positive": "No access points or all with least-privilege policies.",
        "negative": "Excess or permissive policies.",
        "remed": "Audit access point policies; remove unused; deny public; scope principals and prefixes strictly."
    },
    "S3-020": {
        "what": "Object-level SSE on sampled objects.",
        "positive": "aws:kms.",
        "negative": "AES256 or none.",
        "remed": "Re-upload or rewrite objects with SSE-KMS; enforce bucket default encryption and deny non-KMS uploads (S3-006/S3-012)."
    },
}

def s3_info_for_code(code: str) -> Dict[str, str]:
    return S3_CHECK_INFO.get(code, {
        "what": "N/A",
        "positive": "N/A",
        "negative": "N/A",
        "remed": "See recommendation in finding.",
    })

# -------- Report builders per service --------

def build_s3_sheet(wb: Workbook, artifacts: List[Dict[str, Any]]):
    ws = wb.active
    ws.title = "S3"

    # Top header
    ws["A1"] = "LazyAWS Report — S3"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated (UTC): {datetime.now(timezone.utc).isoformat(timespec='seconds')}"
    ws["A3"] = f"Sources: {len(artifacts)} bucket(s)"
    ws_set_col_widths(ws, [36, 10, 50, 50, 42, 84, 84])  # tune widths

    row = 5
    for art in artifacts:
        meta = art.get("meta", {})
        failed = art.get("failed_findings", [])
        trace = art.get("api_trace", [])

        bucket = meta.get("Target", "unknown-bucket")
        region = meta.get("Region", "unknown-region")
        profile = meta.get("Profile", "")
        acc = meta.get("AccountId", "")

        # Section title
        ws.cell(row=row, column=1, value=f"Bucket: {bucket}  |  Region: {region}  |  Account: {acc}  |  Profile: {profile}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        headers = [
            "Check (code & title)",
            "Status",
            "Details",
            "What is tested",
            "Expected (positive / negative)",
            "CLI request (as in console)",
            "CLI response (console output)",
        ]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row, column=c, value=h)
        apply_header_style(ws, f"A{row}:{get_column_letter(len(headers))}{row}")
        row += 1

        start_table_row = row
        if not failed:
            ws.cell(row=row, column=1, value="No failed checks. ✅")
            row += 2
            continue

        for f in failed:
            parsed = parse_check_label(f.get("Check", ""))
            code = parsed["code"]
            title = parsed["title"]
            api_field = parsed["api"]

            svc, op_snake = api_from_check_api(api_field) if api_field else ("", "")
            tr = match_trace_for_api(trace, svc, op_snake) if svc and op_snake else None

            info = s3_info_for_code(code)
            status = f.get("StatusPlain", f.get("Status", ""))  # plain symbol
            details = f.get("Details", "")

            what = info["what"]
            expected = f"POSITIVE: {info['positive']}\nNEGATIVE: {info['negative']}"

            # CLI cmd/output
            cli_cmd = tr.get("cli_cmd") if tr else "N/A"
            cli_out = tr.get("cli_output") if tr else "N/A"

            ws.cell(row=row, column=1, value=f"{code} {title}")
            ws.cell(row=row, column=2, value=status)
            ws.cell(row=row, column=3, value=details)
            ws.cell(row=row, column=4, value=what)

            # expected and remediation (on separate lines)
            remediation = info["remed"]
            ws.cell(row=row, column=5, value=f"{expected}\n\nRemediation:\n{remediation}")
            ws.cell(row=row, column=6, value=cli_cmd)
            ws.cell(row=row, column=7, value=cli_out)

            # Style: wrap all but keep alignment top-left
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            row += 1

        end_table_row = row - 1
        apply_table_borders(ws, start_table_row - 1, end_table_row, 1, 7)
        row += 2  # gap between buckets

def build_generic_sheet(wb: Workbook, name: str, artifacts: List[Dict[str, Any]]):
    ws = wb.create_sheet(title=name.upper())
    ws["A1"] = f"LazyAWS Report — {name.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated (UTC): {datetime.now(timezone.utc).isoformat(timespec='seconds')}"
    ws["A3"] = f"Sources: {len(artifacts)} target(s)"
    ws_set_col_widths(ws, [36, 10, 50, 50, 42, 84, 84])

    row = 5
    if not artifacts:
        ws.cell(row=row, column=1, value="No data.")
        return

    for art in artifacts:
        meta = art.get("meta", {})
        failed = art.get("failed_findings", [])
        trace = art.get("api_trace", [])

        target = meta.get("Target", "unknown")
        region = meta.get("Region", "unknown-region")
        profile = meta.get("Profile", "")
        acc = meta.get("AccountId", "")

        ws.cell(row=row, column=1, value=f"Target: {target}  |  Region: {region}  |  Account: {acc}  |  Profile: {profile}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        headers = [
            "Check (code & title)",
            "Status",
            "Details",
            "What is tested / Expected",
            "Remediation",
            "CLI request",
            "CLI response",
        ]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row, column=c, value=h)
        apply_header_style(ws, f"A{row}:{get_column_letter(len(headers))}{row}")
        row += 1

        start_table_row = row
        if not failed:
            ws.cell(row=row, column=1, value="No failed checks. ✅")
            row += 2
            continue

        for f in failed:
            parsed = parse_check_label(f.get("Check", ""))
            code = parsed["code"]
            title = parsed["title"]
            api_field = parsed["api"]

            svc, op_snake = api_from_check_api(api_field) if api_field else ("", "")
            tr = match_trace_for_api(trace, svc, op_snake) if svc and op_snake else None

            status = f.get("StatusPlain", f.get("Status", ""))
            details = f.get("Details", "")

            # For non-S3 modules (until we define their KB), use generic texts
            what_expected = f"CHECK: {parsed.get('title','')}\nSCOPE: {parsed.get('scope','')}\nEXPECTED: pass (no warnings/errors)."
            remediation = f.get("Recommendation", "Follow service best practices.")

            cli_cmd = tr.get("cli_cmd") if tr else "N/A"
            cli_out = tr.get("cli_output") if tr else "N/A"

            ws.cell(row=row, column=1, value=f"{code} {title}")
            ws.cell(row=row, column=2, value=status)
            ws.cell(row=row, column
